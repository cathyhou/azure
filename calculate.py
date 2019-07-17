from openpyxl import load_workbook

wb = load_workbook(filename = 'results.xlsx')
ws = wb['results']

r = [0, 0, 0, 0, 0]

# calculate accuracy

count = 0


def toIndex(emo):
    emotions = ['HAPPY', 'NEUTRAL', 'SCARED', 'ANGRY', 'SAD']
    for i in range(5):
        if emo==emotions[i]:
            return i
    return 'error'


for i in range(128):
    if ws['D'+str(i+2)].value:
        count = count + 1

for i in range(128):
    row = str(i+2)
    index = toIndex(ws['E'+row].value)
    current = r[index]
    r[index] = current + 1

print(r)
print(count/128.0)