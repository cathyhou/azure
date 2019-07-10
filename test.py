import cognitive_face as CF
import pandas as pd
import sys
from openpyxl import load_workbook
import time

wb = load_workbook(filename = 'results.xlsx')
ws = wb['results']

# Replace with a valid subscription key (keeping the quotes in place).
KEY = '92d758df5cd64cc9a2339c33d21adebd'
CF.Key.set(KEY)

# Replace with your regional Base URL
BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0'
CF.BaseUrl.set(BASE_URL)


def change(emotion):
    current = ['anger', 'contempt', 'disgust', 'fear', 'happiness', 'neutral', 'sadness', 'surprise']
    transform = ['ANGRY', 'CONTEMPT', 'DISGUST', 'SCARED', 'HAPPY', 'NEUTRAL', 'SAD', 'SURPRISED']
    result = ''
    for i in range(8):
        if current[i] == emotion:
            result = transform[i]
            break
    return result


img_url = ''
label = 'SAD'
row=114

faces = CF.face.detect(img_url, attributes='emotion')

#print(faces)
#sys.exit()

emotions = ['anger', 'contempt', 'disgust', 'fear', 'happiness', 'neutral', 'sadness', 'surprise']

anger = 0.000
fear = 0.000
happiness = 0.000
neutral = 0.000
sadness = 0.000

confidence = 0.000
emotion = ''

for emo in emotions:
    if emo == 'n':
        continue
    else:
        e = emo
        emo = faces[0]['faceAttributes']['emotion'][emo]
        if emo > confidence:
            confidence = emo
            emotion = change(e)

ws['B' + str(row)] = emotion
ws['C' + str(row)] = confidence
ws['D' + str(row)] = emotion==label
wb.save('results.xlsx')



