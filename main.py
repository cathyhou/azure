import cognitive_face as CF
import pandas as pd
from openpyxl import load_workbook
import time

wb = load_workbook(filename = 'results.xlsx')
ws = wb['results']
row = 2

KEY = '92d758df5cd64cc9a2339c33d21adebd'
CF.Key.set(KEY)

BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0'
CF.BaseUrl.set(BASE_URL)


def change(emotion):
    current = ['anger', 'contempt', 'disgust', 'fear', 'happiness', 'neutral', 'sadness', 'surprise']
    transform = ['ANGRY', 'CONTEMPT', 'DISGUST', 'SCARED', 'HAPPY', 'NEUTRAL', 'SAD', 'SURPRISED']
    result = ''
    for i in range(8):
        if current[i] == emotion:
            result = transform[i]
            break
    return result


export = pd.read_excel('export.xlsx')

for f in range(len(export)):
    if export['type'].at[f] == 'TEST':  # read export file to determine if image is test image, if not move on
        img_url = export['name'].at[f][60:]  # img_url is path to image
        label = export['label'].at[f]  # label is actual emotion

        if row % 20 == 0 :
            time.sleep(60)  # azure can only do 20images/min

        faces = CF.face.detect(img_url, attributes='emotion')  # run azure api

        if faces == []:  # if no face detected, leave row blank+run with original image later
            ws['A' + str(row)] = img_url
            wb.save('results.xlsx')
            row = row+1
        else :
            emotions = ['anger', 'contempt', 'disgust', 'fear', 'happiness', 'neutral', 'sadness', 'surprise']

            anger = 0.000
            fear = 0.000
            happiness = 0.000
            neutral = 0.000
            sadness = 0.000

            confidence = 0.000
            emotion = ''

            for emo in emotions:  # determine emotion with highest confidence
                if emo == 'n':
                    continue
                else:
                    e = emo
                    emo = faces[0]['faceAttributes']['emotion'][emo]
                    if emo > confidence:
                        confidence = emo
                        emotion = change(e)

            ws['A'+str(row)] = img_url
            ws['B' + str(row)] = emotion
            ws['C' + str(row)] = confidence
            ws['D' + str(row)] = emotion==label  # boolean (for accuracy)
            wb.save('results.xlsx')  # save image file + emotion
            row = row+1
            print(row)


